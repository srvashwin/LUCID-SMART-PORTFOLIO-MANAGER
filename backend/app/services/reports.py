import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import io as io_module
from datetime import date, timedelta
from typing import List, Dict


def generate_excel_report(expenses: List[Dict], monthly_trend: List[Dict], category_breakdown: List[Dict]) -> io_module.BytesIO:
    wb = openpyxl.Workbook()

    # Sheet 1: All Expenses
    ws1 = wb.active
    ws1.title = "Expenses"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    headers = ["Date", "Amount", "Category", "Description", "Merchant", "Use Case"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, exp in enumerate(expenses, 2):
        ws1.cell(row=row, column=1, value=str(exp.get("date", "")))
        ws1.cell(row=row, column=2, value=exp.get("amount", 0))
        ws1.cell(row=row, column=3, value=exp.get("category", ""))
        ws1.cell(row=row, column=4, value=exp.get("description", ""))
        ws1.cell(row=row, column=5, value=exp.get("merchant", ""))
        ws1.cell(row=row, column=6, value=exp.get("use_case", ""))

    # Sheet 2: Monthly Summary
    ws2 = wb.create_sheet("Monthly Summary")
    for col, h in enumerate(["Year", "Month", "Total"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row, t in enumerate(monthly_trend, 2):
        ws2.cell(row=row, column=1, value=t.get("year", ""))
        ws2.cell(row=row, column=2, value=t.get("month", ""))
        ws2.cell(row=row, column=3, value=t.get("total", 0))

    # Sheet 3: Category Breakdown
    ws3 = wb.create_sheet("Category Breakdown")
    for col, h in enumerate(["Category", "Total", "Count"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row, c in enumerate(category_breakdown, 2):
        ws3.cell(row=row, column=1, value=c.get("category", ""))
        ws3.cell(row=row, column=2, value=c.get("total", 0))
        ws3.cell(row=row, column=3, value=c.get("count", 0))

    # Auto-adjust column widths
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    output = io_module.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_expense_chart(monthly_trend: List[Dict]) -> io_module.BytesIO:
    fig, ax = plt.subplots(figsize=(10, 5))
    months = [f"{t['year']}-{str(t['month']).zfill(2)}" for t in monthly_trend]
    totals = [t["total"] for t in monthly_trend]

    ax.plot(months, totals, marker="o", linewidth=2, color="#2563EB")
    ax.fill_between(range(len(months)), totals, alpha=0.1, color="#2563EB")
    ax.set_xlabel("Month")
    ax.set_ylabel("Total Spending ($)")
    ax.set_title("Monthly Spending Trend")
    ax.tick_params(axis="x", rotation=45)
    ax.grid(True, alpha=0.3)

    plt.tight_layout()
    buf = io_module.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    buf.seek(0)
    plt.close(fig)
    return buf


def generate_category_chart(category_breakdown: List[Dict]) -> io_module.BytesIO:
    fig, ax = plt.subplots(figsize=(8, 8))
    labels = [c["category"] for c in category_breakdown if c["total"] > 0]
    sizes = [c["total"] for c in category_breakdown if c["total"] > 0]
    colors = plt.cm.Blues([0.3 + 0.7 * i / max(len(sizes), 1) for i in range(len(sizes))])

    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, autopct="%1.1f%%",
        colors=colors, startangle=90, textprops={"fontsize": 9}
    )
    ax.set_title("Spending by Category")

    plt.tight_layout()
    buf = io_module.BytesIO()
    fig.savefig(buf, format="png", dpi=150)
    buf.seek(0)
    plt.close(fig)
    return buf
