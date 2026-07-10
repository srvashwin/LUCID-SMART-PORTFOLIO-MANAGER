import openpyxl
import io


def test_generate_excel_report_has_three_sheets():
    from app.services.reports import generate_excel_report
    buf = generate_excel_report([], [], [])
    wb = openpyxl.load_workbook(buf)
    assert wb.sheetnames == ["Expenses", "Monthly Summary", "Category Breakdown"]


def test_generate_excel_report_expenses():
    from app.services.reports import generate_excel_report
    expenses = [{"date": "2024-03-15", "amount": 150.0, "category": "Food", "description": "Lunch", "merchant": "Cafe", "use_case": ""}]
    buf = generate_excel_report(expenses, [], [])
    wb = openpyxl.load_workbook(buf)
    ws = wb["Expenses"]
    assert ws.cell(2, 1).value == "2024-03-15"
    assert ws.cell(2, 2).value == 150.0
    assert ws.cell(2, 3).value == "Food"


def test_generate_excel_report_monthly_trend():
    from app.services.reports import generate_excel_report
    trend = [{"year": 2024, "month": 3, "total": 5000}]
    buf = generate_excel_report([], trend, [])
    wb = openpyxl.load_workbook(buf)
    ws = wb["Monthly Summary"]
    assert ws.cell(2, 1).value == 2024
    assert ws.cell(2, 2).value == 3
    assert ws.cell(2, 3).value == 5000


def test_generate_excel_report_category_breakdown():
    from app.services.reports import generate_excel_report
    cats = [{"category": "Food", "total": 2000, "count": 10}]
    buf = generate_excel_report([], [], cats)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Category Breakdown"]
    assert ws.cell(2, 1).value == "Food"
    assert ws.cell(2, 2).value == 2000
    assert ws.cell(2, 3).value == 10


def test_generate_excel_report_column_headers():
    from app.services.reports import generate_excel_report
    buf = generate_excel_report([], [], [])
    wb = openpyxl.load_workbook(buf)
    ws = wb["Expenses"]
    assert ws.cell(1, 1).value == "Date"
    assert ws.cell(1, 2).value == "Amount"
    assert ws.cell(1, 3).value == "Category"
    assert ws.cell(1, 4).value == "Description"
    assert ws.cell(1, 5).value == "Merchant"
    assert ws.cell(1, 6).value == "Use Case"


def test_generate_excel_returns_bytesio():
    from app.services.reports import generate_excel_report
    buf = generate_excel_report([], [], [])
    assert isinstance(buf, io.BytesIO)
    assert buf.getvalue()[:2] == b"PK"


def test_generate_expense_chart_returns_png():
    from app.services.reports import generate_expense_chart
    trend = [{"year": 2024, "month": 1, "total": 5000}, {"year": 2024, "month": 2, "total": 4500}]
    buf = generate_expense_chart(trend)
    assert buf.getvalue()[:8] == b"\x89PNG\r\n\x1a\n"


def test_generate_expense_chart_single_point():
    from app.services.reports import generate_expense_chart
    trend = [{"year": 2024, "month": 1, "total": 5000}]
    buf = generate_expense_chart(trend)
    assert buf.getvalue()[:8] == b"\x89PNG\r\n\x1a\n"


def test_generate_category_chart_returns_png():
    from app.services.reports import generate_category_chart
    cats = [{"category": "Food", "total": 2000}, {"category": "Transport", "total": 500}]
    buf = generate_category_chart(cats)
    assert buf.getvalue()[:8] == b"\x89PNG\r\n\x1a\n"


def test_generate_category_chart_empty():
    from app.services.reports import generate_category_chart
    buf = generate_category_chart([])
    assert buf.getvalue()[:8] == b"\x89PNG\r\n\x1a\n"


def test_generate_category_chart_all_zero_filtered():
    from app.services.reports import generate_category_chart
    cats = [{"category": "Food", "total": 0}, {"category": "Transport", "total": 0}]
    buf = generate_category_chart(cats)
    assert buf.getvalue()[:8] == b"\x89PNG\r\n\x1a\n"
